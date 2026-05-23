#!/usr/bin/env python3
"""
Fidelavis Lifestyle Generator v2.1
Génère wallets premium + mockups DM pour restaurants brunch/gastro/coffee
Usage : python3 generate.py [chemin_excel]
"""

import base64
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Optional
from urllib.request import Request, urlopen

import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

EXCEL_PATH = Path("/Users/jcv/Documents/fichier resto/brunch/brunch11/lebon.xlsx")
REPO_DIR   = Path(__file__).resolve().parent.parent
OUTPUT_DIR = REPO_DIR / "lifestyle" / "restaurants"
DEMO_BASE_URL = "https://app.cartefidelavis.com/demo"
CURRENT_MONTH = date.today().strftime("%B %Y")

# ── Style detection ───────────────────────────────────────────────────────────

STYLE_SIGNALS = {
    "brunch": [
        "brunch", "breakfast", "petit-déjeuner", "lunch dinner", "mimosa",
        "avocado", "pancake", "waffle", "gaufre", "granola", "sunday",
        "brunch lunch", "weekend brunch",
    ],
    "gastro": [
        "gastronomique", "bistronomie", "fine dining", "étoilé", "michelin",
        "chef", "haute cuisine", "gourmet", "dégustation", "tasting",
        "gastro", "gastronomic",
    ],
    "coffee": [
        "coffee", "coffee shop", "barista", "espresso", "specialty coffee",
        "torréfacteur", "roaster", "cold brew", "brûlerie", "café de spécialité",
        "third wave",
    ],
}

CONTENT = {
    "brunch": {
        "taglines": [
            "Instagrammable & gourmand",
            "Le spot brunch du quartier",
            "Brunch, partage & bonne humeur",
        ],
        "offers": [
            ("Dessert du chef", "Sur présentation de votre code en caisse"),
            ("Mimosa à l'arrivée", "Sur réservation, le week-end"),
            ("Granola bowl maison", "Pour tout brunch à partir de 2 personnes"),
        ],
        "subtitle": "BRUNCH CLUB",
        "share_label": "Partager ce brunch",
        "share_sub": "Envoyez cette adresse à vos amis et partagez un bon moment",
    },
    "gastro": {
        "taglines": [
            "L'art de recevoir à la française",
            "Une expérience gastronomique",
            "La table des amateurs de bonne cuisine",
        ],
        "offers": [
            ("Cocktail signature", "Sur présentation de votre code en caisse"),
            ("Mignardises", "En fin de repas, pour les membres"),
            ("Expérience signature", "Pour toute table de 2 personnes"),
        ],
        "subtitle": "RESTAURANT",
        "share_label": "Partager cette adresse",
        "share_sub": "Recommandez cette table à vos proches",
    },
    "coffee": {
        "taglines": [
            "Specialty coffee & daily rituals",
            "L'art du café de spécialité",
            "Le café de vos rituels du matin",
        ],
        "offers": [
            ("Café signature", "Sur présentation de votre code en caisse"),
            ("Viennoiserie", "Avec tout café de spécialité"),
            ("Dégustation découverte", "Sélection du mois"),
        ],
        "subtitle": "COFFEE SHOP",
        "share_label": "Partager ce café",
        "share_sub": "Faites découvrir votre adresse préférée",
    },
}


def detect_style(name: str, cat1: str, cat2: str) -> str:
    text = " ".join(filter(None, [name, cat1, cat2])).lower()
    scores = {s: sum(1 for kw in kws if kw in text) for s, kws in STYLE_SIGNALS.items()}
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else "brunch"


def slugify(name: str) -> str:
    s = name.lower()
    for src, dst in [("àáâã", "a"), ("éèêë", "e"), ("îï", "i"), ("ôö", "o"), ("ùúûü", "u"), ("ç", "c")]:
        for c in src:
            s = s.replace(c, dst)
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")[:40]


def fetch_url_b64(url: str, timeout: int = 8) -> Optional[str]:
    if not url:
        return None
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=timeout) as resp:
            data = resp.read()
        return base64.b64encode(data).decode()
    except Exception:
        return None


def initials(name: str) -> str:
    words = name.split()[:2]
    return "".join(w[0].upper() for w in words if w)


def extract_instagram_handle(url: str) -> str:
    if not url:
        return ""
    m = re.search(r"instagram\.com/([^/?#\s]+)", url)
    if m:
        h = m.group(1).rstrip("/")
        return f"@{h}" if h else ""
    return ""


def header_name(name: str) -> str:
    stop = {"brunch", "lunch", "dinner", "restaurant", "paris", "coffee", "shop",
            "café", "cafe", "bistro", "bar", "de", "la", "le", "les", "du"}
    words = name.split()
    keep = [w for w in words if w.lower() not in stop][:3] or words[:2]
    return " ".join(keep).upper()[:22]


# ── Instagram hero image scraping ─────────────────────────────────────────────

def upgrade_cdn_url(url: str) -> str:
    """Upgrade Instagram CDN thumbnail URL to highest available resolution."""
    # Profile pics: t51.2885-19, s100x100 → s640x640
    url = re.sub(r"(t51\.2885-19)/[^/]*/", r"\1/", url)
    url = re.sub(r"s\d+x\d+/", "s640x640/", url)
    url = re.sub(r"_s\.jpg", "_n.jpg", url)
    return url


def scrape_instagram_hero(instagram_url: str) -> Optional[str]:
    """
    Scrape the first feed image from an Instagram profile.
    Returns base64-encoded image data, or None on failure.
    Strategy:
      1. Open profile page with Playwright (no login, public accounts)
      2. Find first <img> in the feed grid (not profile pic: t51.2885-19 is profile,
         t51.2885-15 / t51.2885-34 are post thumbnails)
      3. Upgrade to higher-res CDN URL
      4. Fallback: upgrade profile picture URL to s640x640
    """
    if not instagram_url:
        return None
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("  ⚠  playwright non installé — pas de scraping Instagram")
        return None

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) "
                    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1"
                ),
                viewport={"width": 390, "height": 844},
                locale="fr-FR",
            )
            page = ctx.new_page()
            page.goto(instagram_url, wait_until="networkidle", timeout=20000)
            page.wait_for_timeout(2500)

            # Collect all img srcs
            srcs = page.eval_on_selector_all("img[src]", "imgs => imgs.map(i => i.src)")

            post_url = None
            profile_url = None

            for src in srcs:
                if not src or "blob:" in src:
                    continue
                if "t51.2885-19" in src and not profile_url:
                    profile_url = src
                elif ("t51.2885-15" in src or "t51.2885-34" in src or "cdninstagram" in src) and not post_url:
                    if "150x150" not in src or "s150x150" not in src:
                        post_url = src

            browser.close()

            # Prefer first post image, fallback to higher-res profile pic
            hero_url = post_url or profile_url
            if not hero_url:
                return None

            hero_url = upgrade_cdn_url(hero_url)
            return fetch_url_b64(hero_url, timeout=10)

    except Exception as e:
        print(f"  ⚠  Instagram scraping échoué ({e})")
        return None


# ── BRUNCH V1 template ────────────────────────────────────────────────────────

def render_wallet_brunch(d: dict) -> str:
    if d.get("hero_b64"):
        hero_src = f'data:image/jpeg;base64,{d["hero_b64"]}'
        hero_bg_style = f'background-image:url("{hero_src}");background-size:cover;background-position:center;'
    elif d.get("logo_b64"):
        hero_src = f'data:image/jpeg;base64,{d["logo_b64"]}'
        hero_bg_style = f'background-image:url("{hero_src}");background-size:cover;background-position:center;'
    else:
        hero_bg_style = "background:linear-gradient(135deg,#F9C8D4 0%,#F0A8B8 50%,#E890A8 100%);"

    if d.get("logo_b64"):
        logo_img = f'<img src="data:image/jpeg;base64,{d["logo_b64"]}" alt="logo">'
    else:
        logo_img = f'<div class="logo-init">{d["initials"]}</div>'

    insta_handle = extract_instagram_handle(d.get("instagram_url", ""))
    h_name = header_name(d["name"])

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=390, initial-scale=1.0">
<title>{d['name']} × Fidelavis</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400;0,700;1,400&family=Dancing+Script:wght@500;600;700&family=Lato:wght@300;400;700&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --rose:#D4637A;--rose-mid:#E8809A;--rose-light:#F5C2C8;
  --rose-pale:#FFF0F3;--rose-bg:#FDE8EC;
  --cream:#FDFAF8;--text:#2D1015;--muted:#8A6570;--gold:#C9962A;
}}
body{{font-family:'Lato',sans-serif;background:var(--rose-pale);
  max-width:390px;margin:0 auto;overflow-x:hidden;color:var(--text)}}
.hero-wrap{{position:relative;height:360px;overflow:hidden}}
.hero-bg{{
  position:absolute;inset:-30px;
  {hero_bg_style}
  filter:blur(28px) saturate(1.5) brightness(1.05);
  transform:scale(1.15);
}}
.hero-overlay{{
  position:absolute;inset:0;
  background:linear-gradient(180deg,rgba(255,230,235,.72) 0%,rgba(250,195,210,.82) 60%,rgba(245,175,195,.88) 100%);
}}
.status-bar{{
  position:absolute;top:0;left:0;right:0;z-index:10;
  display:flex;justify-content:space-between;align-items:center;
  padding:14px 20px 0;font-size:15px;font-weight:700;color:var(--text);
}}
.status-icons{{display:flex;align-items:center;gap:4px;font-size:12px}}
.hero-logo-wrap{{
  position:absolute;top:42px;left:20px;z-index:8;
  width:74px;height:74px;border-radius:50%;overflow:hidden;
  border:3px solid rgba(255,255,255,.9);
  box-shadow:0 4px 20px rgba(212,99,122,.25);
  background:var(--rose-light);
  display:flex;align-items:center;justify-content:center;
}}
.hero-logo-wrap img{{width:100%;height:100%;object-fit:cover}}
.logo-init{{font-family:'Playfair Display',serif;font-size:22px;font-weight:700;color:#fff;}}
.hero-bell{{
  position:absolute;top:44px;right:20px;z-index:8;
  width:40px;height:40px;border-radius:50%;
  background:rgba(255,255,255,.8);backdrop-filter:blur(8px);
  display:flex;align-items:center;justify-content:center;
  font-size:18px;box-shadow:0 2px 10px rgba(0,0,0,.08);cursor:pointer;
}}
.hero-center{{
  position:absolute;top:50%;left:50%;
  transform:translate(-50%,-52%);text-align:center;z-index:5;width:220px;
}}
.hero-sun{{font-size:20px;color:var(--gold);margin-bottom:6px;line-height:1}}
.hero-name{{
  font-family:'Playfair Display',serif;font-size:26px;font-weight:700;
  color:var(--text);letter-spacing:2px;text-transform:uppercase;line-height:1.1;
  text-shadow:0 1px 8px rgba(255,255,255,.6);
}}
.hero-subtitle{{
  font-size:11px;font-weight:700;letter-spacing:5px;
  color:var(--rose);text-transform:uppercase;margin-top:5px;
}}
.hero-tagline{{
  font-family:'Dancing Script',cursive;font-size:17px;
  color:var(--muted);margin-top:7px;line-height:1.2;
}}
.profile-card{{
  position:absolute;bottom:18px;left:16px;z-index:8;
  background:rgba(255,255,255,.9);backdrop-filter:blur(12px);
  border-radius:18px;padding:13px 18px 13px 16px;
  box-shadow:0 8px 32px rgba(212,99,122,.18);min-width:180px;
}}
.profile-hello{{font-family:'Dancing Script',cursive;font-size:18px;color:var(--rose);line-height:1;}}
.profile-name{{font-family:'Playfair Display',serif;font-size:28px;font-weight:700;color:var(--text);line-height:1.1;margin-top:1px;}}
.profile-since{{font-size:12px;color:var(--muted);margin-top:5px;display:flex;align-items:center;gap:4px;}}
.crown{{color:var(--gold)}}
.content{{padding:16px 16px 100px}}
.offer-card{{
  border-radius:22px;overflow:hidden;position:relative;
  height:230px;margin-bottom:14px;
  background:linear-gradient(145deg,#4A1525 0%,#7A2840 35%,#B04060 70%,#C85870 100%);
}}
.offer-card-deco{{
  position:absolute;inset:0;
  background:
    radial-gradient(ellipse 65% 50% at 75% 35%,rgba(230,130,80,.35),transparent),
    radial-gradient(ellipse 45% 60% at 15% 65%,rgba(180,50,80,.3),transparent);
}}
.offer-body{{position:relative;z-index:2;padding:20px;height:100%;display:flex;flex-direction:column;justify-content:flex-end;}}
.offer-pill{{
  display:inline-block;background:var(--rose);color:#fff;
  font-size:9px;font-weight:700;letter-spacing:1.8px;
  text-transform:uppercase;padding:5px 13px;border-radius:30px;
  margin-bottom:10px;align-self:flex-start;
}}
.offer-title-text{{font-family:'Playfair Display',serif;font-size:26px;font-weight:700;color:#fff;line-height:1.1;}}
.offer-italic{{font-family:'Dancing Script',cursive;font-size:28px;font-weight:700;color:var(--rose-light);line-height:1;margin-top:2px;}}
.offer-desc{{font-size:12px;color:rgba(255,255,255,.65);margin-top:7px;line-height:1.4;}}
.offer-badge{{
  position:absolute;right:16px;bottom:16px;z-index:3;
  width:76px;height:76px;border-radius:50%;
  background:rgba(212,99,122,.88);border:2px solid rgba(255,255,255,.4);
  display:flex;align-items:center;justify-content:center;
  text-align:center;color:#fff;font-size:8.5px;font-weight:700;
  letter-spacing:.6px;text-transform:uppercase;line-height:1.25;padding:10px;
}}
.actions-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px;}}
.action-card{{
  background:#fff;border-radius:18px;padding:16px 14px 14px;
  box-shadow:0 2px 14px rgba(212,99,122,.07);
  text-decoration:none;color:inherit;display:flex;flex-direction:column;gap:3px;position:relative;
}}
.action-icon-circle{{
  width:44px;height:44px;border-radius:50%;
  background:linear-gradient(135deg,#FDEAED,#F7D0D6);
  display:flex;align-items:center;justify-content:center;
  font-size:20px;margin-bottom:8px;
}}
.action-label{{font-size:13px;font-weight:700;color:var(--text);line-height:1.2}}
.action-desc{{font-size:11px;color:var(--muted);line-height:1.3}}
.action-desc.pink{{color:var(--rose)}}
.action-arrow{{position:absolute;right:12px;top:50%;transform:translateY(-50%);color:var(--rose-light);font-size:20px;font-weight:300;}}
.share-banner{{
  border-radius:22px;overflow:hidden;
  background:linear-gradient(120deg,#D4637A 0%,#C04868 60%,#A83858 100%);
  padding:20px 18px;display:flex;align-items:center;gap:14px;
  position:relative;text-decoration:none;color:inherit;min-height:96px;
}}
.share-banner::after{{content:'';position:absolute;right:0;top:0;bottom:0;width:110px;background:linear-gradient(90deg,transparent,rgba(180,40,60,.25));border-radius:0 22px 22px 0;}}
.share-env{{width:52px;height:52px;flex-shrink:0;border-radius:50%;background:rgba(255,255,255,.2);display:flex;align-items:center;justify-content:center;font-size:24px;z-index:1;}}
.share-text{{z-index:1;flex:1}}
.share-title{{font-family:'Dancing Script',cursive;font-size:24px;font-weight:700;color:#fff;line-height:1;}}
.share-sub{{font-size:11px;color:rgba(255,255,255,.72);margin-top:4px;line-height:1.35;}}
.share-arrow{{color:rgba(255,255,255,.55);font-size:24px;z-index:1;}}
.bottom-nav{{
  position:fixed;bottom:0;left:50%;transform:translateX(-50%);
  width:100%;max-width:390px;
  background:rgba(255,240,243,.96);backdrop-filter:blur(16px);
  border-top:1px solid rgba(212,99,122,.12);
  padding:10px 0 24px;display:flex;z-index:20;
}}
.nav-tab{{flex:1;display:flex;flex-direction:column;align-items:center;gap:3px;color:var(--muted);font-size:12px;text-decoration:none;transition:color .2s;}}
.nav-tab.active{{color:var(--rose)}}
.nav-icon{{font-size:22px;line-height:1}}
</style>
</head>
<body>
<div class="hero-wrap">
  <div class="hero-bg"></div>
  <div class="hero-overlay"></div>
  <div class="status-bar">
    <span>9:41</span>
    <div class="status-icons">
      <svg width="17" height="12" viewBox="0 0 17 12" fill="none"><rect x="0" y="3" width="3" height="9" rx="1" fill="#2D1015"/><rect x="4.5" y="2" width="3" height="10" rx="1" fill="#2D1015"/><rect x="9" y="0" width="3" height="12" rx="1" fill="#2D1015"/><rect x="13.5" y="0" width="3" height="12" rx="1" fill="rgba(45,16,21,.3)"/></svg>
      <svg width="25" height="12" viewBox="0 0 25 12" fill="none"><rect x="0" y="1" width="22" height="10" rx="3" stroke="#2D1015" stroke-width="1"/><rect x="1.5" y="2.5" width="16" height="7" rx="2" fill="#2D1015"/><path d="M23 4v4a2 2 0 000-4z" fill="#2D1015"/></svg>
    </div>
  </div>
  <div class="hero-logo-wrap">{logo_img}</div>
  <div class="hero-bell">🔔</div>
  <div class="hero-center">
    <div class="hero-sun">✦</div>
    <div class="hero-name">{h_name}</div>
    <div class="hero-subtitle">{d['content']['subtitle']}</div>
    <div class="hero-tagline">{d['tagline']}</div>
  </div>
  <div class="profile-card">
    <div class="profile-hello">Bonjour ♡</div>
    <div class="profile-name">Marie ✨</div>
    <div class="profile-since"><span class="crown">👑</span> Membre depuis {CURRENT_MONTH}</div>
  </div>
</div>
<div class="content">
  <div class="offer-card">
    <div class="offer-card-deco"></div>
    <div class="offer-body">
      <div class="offer-pill">Avantage privé du mois</div>
      <div class="offer-title-text">{d['offer_title']}</div>
      <div class="offer-italic">offert ♡</div>
      <div class="offer-desc">{d['offer_sub']}</div>
    </div>
    <div class="offer-badge">RÉSERVÉ<br>À NOS<br>MEMBRES</div>
  </div>
  <div class="actions-grid">
    <a href="#" class="action-card">
      <div class="action-icon-circle">✨</div>
      <div class="action-label">✨ L'offre du moment</div>
      <div class="action-desc">Découvrez l'offre exclusive</div>
      <span class="action-arrow">›</span>
    </a>
    <a href="{d.get('website') or '#'}" class="action-card">
      <div class="action-icon-circle">📅</div>
      <div class="action-label">📅 Réserver</div>
      <div class="action-desc">Réservez votre table</div>
      <span class="action-arrow">›</span>
    </a>
    <a href="{d.get('instagram_url') or '#'}" class="action-card">
      <div class="action-icon-circle">📸</div>
      <div class="action-label">📸 Instagram</div>
      <div class="action-desc">Suivez-nous</div>
      <div class="action-desc pink">{insta_handle}</div>
      <span class="action-arrow">›</span>
    </a>
    <a href="{d.get('review_url') or '#'}" class="action-card">
      <div class="action-icon-circle">⭐</div>
      <div class="action-label">⭐ Avis Google</div>
      <div class="action-desc">Partagez votre avis</div>
      <span class="action-arrow">›</span>
    </a>
  </div>
  <a href="#" class="share-banner">
    <div class="share-env">💌</div>
    <div class="share-text">
      <div class="share-title">💌 {d['content']['share_label']}</div>
      <div class="share-sub">{d['content']['share_sub']}</div>
    </div>
    <span class="share-arrow">›</span>
  </a>
</div>
<div class="bottom-nav">
  <a href="#" class="nav-tab active"><span class="nav-icon">👑</span>Club</a>
  <a href="#" class="nav-tab"><span class="nav-icon">👤</span>Compte</a>
</div>
</body>
</html>"""


# ── GASTRO V1 template ─────────────────────────────────────────────────────────

def render_wallet_gastro(d: dict) -> str:
    if d.get("hero_b64"):
        hero_bg_style = f'background-image:url("data:image/jpeg;base64,{d["hero_b64"]}");background-size:cover;background-position:center;'
    elif d.get("logo_b64"):
        hero_bg_style = f'background-image:url("data:image/jpeg;base64,{d["logo_b64"]}");background-size:cover;background-position:center;'
    else:
        hero_bg_style = "background:linear-gradient(135deg,#1A1200 0%,#2D2000 100%);"

    logo_img = (f'<img src="data:image/jpeg;base64,{d["logo_b64"]}" alt="logo">'
                if d.get("logo_b64") else f'<div class="logo-init">{d["initials"]}</div>')
    insta_handle = extract_instagram_handle(d.get("instagram_url", ""))
    h_name = header_name(d["name"])

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=390, initial-scale=1.0">
<title>{d['name']} × Fidelavis</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,600;1,400&family=Dancing+Script:wght@600;700&family=Montserrat:wght@300;400;600&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{--gold:#C9A84C;--gold-light:#E8D5A3;--black:#0D0D0D;--card:#1A1A1A;--text:#F5F0EB;--muted:#8A8070;}}
body{{font-family:'Montserrat',sans-serif;background:var(--black);max-width:390px;margin:0 auto;overflow-x:hidden;color:var(--text)}}
.hero-wrap{{position:relative;height:340px;overflow:hidden}}
.hero-bg{{position:absolute;inset:-30px;{hero_bg_style}filter:blur(25px) saturate(1.2) brightness(.55);transform:scale(1.15);}}
.hero-overlay{{position:absolute;inset:0;background:linear-gradient(180deg,rgba(10,8,0,.55) 0%,rgba(20,15,0,.7) 100%);}}
.status-bar{{position:absolute;top:0;left:0;right:0;z-index:10;display:flex;justify-content:space-between;align-items:center;padding:14px 20px 0;font-size:15px;font-weight:600;color:var(--text);}}
.hero-logo-wrap{{position:absolute;top:42px;left:20px;z-index:8;width:70px;height:70px;border-radius:50%;overflow:hidden;border:1px solid rgba(201,168,76,.6);background:var(--card);display:flex;align-items:center;justify-content:center;}}
.hero-logo-wrap img{{width:100%;height:100%;object-fit:cover}}
.logo-init{{font-family:'Cormorant Garamond',serif;font-size:22px;color:var(--gold);}}
.hero-bell{{position:absolute;top:44px;right:20px;z-index:8;width:40px;height:40px;border-radius:4px;background:rgba(201,168,76,.12);border:1px solid rgba(201,168,76,.3);display:flex;align-items:center;justify-content:center;font-size:18px;}}
.hero-center{{position:absolute;top:50%;left:50%;transform:translate(-50%,-52%);text-align:center;z-index:5;width:240px;}}
.hero-sun{{font-size:18px;color:var(--gold);margin-bottom:8px}}
.hero-name{{font-family:'Cormorant Garamond',serif;font-size:28px;font-weight:400;color:var(--text);letter-spacing:4px;text-transform:uppercase;line-height:1.1;}}
.hero-subtitle{{font-size:10px;font-weight:600;letter-spacing:6px;color:var(--gold);text-transform:uppercase;margin-top:6px;}}
.hero-tagline{{font-family:'Dancing Script',cursive;font-size:16px;color:rgba(255,255,255,.55);margin-top:8px;}}
.profile-card{{position:absolute;bottom:18px;left:16px;z-index:8;background:rgba(26,26,26,.92);backdrop-filter:blur(12px);border:1px solid rgba(201,168,76,.2);border-radius:4px;padding:14px 18px;min-width:180px;box-shadow:0 8px 32px rgba(0,0,0,.4);}}
.profile-hello{{font-family:'Dancing Script',cursive;font-size:18px;color:var(--gold)}}
.profile-name{{font-family:'Cormorant Garamond',serif;font-size:28px;color:var(--text);line-height:1.1;margin-top:1px;}}
.profile-since{{font-size:11px;color:var(--muted);margin-top:5px}}
.content{{padding:16px 16px 100px;background:var(--black)}}
.offer-card{{border-radius:4px;overflow:hidden;position:relative;height:220px;margin-bottom:14px;background:linear-gradient(145deg,#0D0A00 0%,#1A1400 40%,#2A2000 100%);border:1px solid rgba(201,168,76,.15);}}
.offer-body{{position:relative;z-index:2;padding:20px;height:100%;display:flex;flex-direction:column;justify-content:flex-end;}}
.offer-pill{{display:inline-block;border:1px solid var(--gold);color:var(--gold);font-size:9px;font-weight:600;letter-spacing:2px;text-transform:uppercase;padding:4px 12px;border-radius:2px;margin-bottom:10px;align-self:flex-start;}}
.offer-title-text{{font-family:'Cormorant Garamond',serif;font-size:28px;color:var(--text);line-height:1.1;}}
.offer-italic{{font-family:'Dancing Script',cursive;font-size:26px;color:var(--gold);line-height:1;margin-top:2px;}}
.offer-desc{{font-size:12px;color:rgba(255,255,255,.5);margin-top:7px}}
.offer-badge{{position:absolute;right:16px;bottom:16px;z-index:3;width:76px;height:76px;border-radius:50%;background:rgba(201,168,76,.15);border:1px solid rgba(201,168,76,.5);display:flex;align-items:center;justify-content:center;text-align:center;color:var(--gold);font-size:8px;font-weight:600;letter-spacing:.8px;text-transform:uppercase;line-height:1.25;padding:10px;}}
.actions-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}}
.action-card{{background:var(--card);border-radius:4px;padding:16px 14px 14px;border:1px solid rgba(201,168,76,.1);text-decoration:none;color:inherit;display:flex;flex-direction:column;gap:3px;position:relative;}}
.action-icon-circle{{width:42px;height:42px;border-radius:50%;background:rgba(201,168,76,.1);border:1px solid rgba(201,168,76,.2);display:flex;align-items:center;justify-content:center;font-size:18px;margin-bottom:8px;}}
.action-label{{font-size:13px;font-weight:600;color:var(--text);line-height:1.2}}
.action-desc{{font-size:11px;color:var(--muted);line-height:1.3}}
.action-desc.gold{{color:var(--gold)}}
.action-arrow{{position:absolute;right:12px;top:50%;transform:translateY(-50%);color:rgba(201,168,76,.3);font-size:20px;}}
.share-banner{{border-radius:4px;overflow:hidden;background:linear-gradient(120deg,#1A1400 0%,#2A2000 100%);border:1px solid rgba(201,168,76,.25);padding:20px 18px;display:flex;align-items:center;gap:14px;text-decoration:none;color:inherit;min-height:96px;}}
.share-env{{width:50px;height:50px;flex-shrink:0;border-radius:50%;background:rgba(201,168,76,.12);border:1px solid rgba(201,168,76,.3);display:flex;align-items:center;justify-content:center;font-size:22px;}}
.share-text{{flex:1}}
.share-title{{font-family:'Dancing Script',cursive;font-size:22px;color:var(--gold);font-weight:700;}}
.share-sub{{font-size:11px;color:var(--muted);margin-top:4px;line-height:1.35}}
.share-arrow{{color:rgba(201,168,76,.4);font-size:22px}}
.bottom-nav{{position:fixed;bottom:0;left:50%;transform:translateX(-50%);width:100%;max-width:390px;background:rgba(13,13,13,.97);backdrop-filter:blur(16px);border-top:1px solid rgba(201,168,76,.12);padding:10px 0 24px;display:flex;z-index:20;}}
.nav-tab{{flex:1;display:flex;flex-direction:column;align-items:center;gap:3px;color:var(--muted);font-size:11px;letter-spacing:.5px;text-decoration:none;}}
.nav-tab.active{{color:var(--gold)}}
.nav-icon{{font-size:20px;line-height:1}}
</style>
</head>
<body>
<div class="hero-wrap">
  <div class="hero-bg"></div><div class="hero-overlay"></div>
  <div class="status-bar"><span>9:41</span><span>▪▪▪ 🔋</span></div>
  <div class="hero-logo-wrap">{logo_img}</div>
  <div class="hero-bell">🔔</div>
  <div class="hero-center">
    <div class="hero-sun">✦</div>
    <div class="hero-name">{h_name}</div>
    <div class="hero-subtitle">{d['content']['subtitle']}</div>
    <div class="hero-tagline">{d['tagline']}</div>
  </div>
  <div class="profile-card">
    <div class="profile-hello">Bonsoir ♡</div>
    <div class="profile-name">Marie</div>
    <div class="profile-since">✦ Membre depuis {CURRENT_MONTH}</div>
  </div>
</div>
<div class="content">
  <div class="offer-card">
    <div class="offer-body">
      <div class="offer-pill">Privilège du mois</div>
      <div class="offer-title-text">{d['offer_title']}</div>
      <div class="offer-italic">offert ✦</div>
      <div class="offer-desc">{d['offer_sub']}</div>
    </div>
    <div class="offer-badge">RÉSERVÉ<br>AUX<br>MEMBRES</div>
  </div>
  <div class="actions-grid">
    <a href="#" class="action-card"><div class="action-icon-circle">✦</div><div class="action-label">L'offre du moment</div><div class="action-desc">Votre avantage exclusif</div><span class="action-arrow">›</span></a>
    <a href="{d.get('website') or '#'}" class="action-card"><div class="action-icon-circle">📅</div><div class="action-label">Réserver</div><div class="action-desc">Réservez votre table</div><span class="action-arrow">›</span></a>
    <a href="{d.get('instagram_url') or '#'}" class="action-card"><div class="action-icon-circle">📸</div><div class="action-label">Instagram</div><div class="action-desc">Découvrir l'univers</div><div class="action-desc gold">{insta_handle}</div><span class="action-arrow">›</span></a>
    <a href="{d.get('review_url') or '#'}" class="action-card"><div class="action-icon-circle">⭐</div><div class="action-label">Avis Google</div><div class="action-desc">Partager votre avis</div><span class="action-arrow">›</span></a>
  </div>
  <a href="#" class="share-banner">
    <div class="share-env">💌</div>
    <div class="share-text"><div class="share-title">{d['content']['share_label']}</div><div class="share-sub">{d['content']['share_sub']}</div></div>
    <span class="share-arrow">›</span>
  </a>
</div>
<div class="bottom-nav">
  <a href="#" class="nav-tab active"><span class="nav-icon">◆</span>Club</a>
  <a href="#" class="nav-tab"><span class="nav-icon">👤</span>Compte</a>
</div>
</body>
</html>"""


# ── COFFEE V1 template ────────────────────────────────────────────────────────

def render_wallet_coffee(d: dict) -> str:
    if d.get("hero_b64"):
        hero_bg_style = f'background-image:url("data:image/jpeg;base64,{d["hero_b64"]}");background-size:cover;background-position:center;'
    elif d.get("logo_b64"):
        hero_bg_style = f'background-image:url("data:image/jpeg;base64,{d["logo_b64"]}");background-size:cover;background-position:center;'
    else:
        hero_bg_style = "background:linear-gradient(135deg,#C4A882 0%,#8B7355 100%);"

    logo_img = (f'<img src="data:image/jpeg;base64,{d["logo_b64"]}" alt="logo">'
                if d.get("logo_b64") else f'<div class="logo-init">{d["initials"]}</div>')
    insta_handle = extract_instagram_handle(d.get("instagram_url", ""))
    h_name = header_name(d["name"])

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=390, initial-scale=1.0">
<title>{d['name']} × Fidelavis</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=Dancing+Script:wght@600;700&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{--coffee:#8B7355;--latte:#C4A882;--cream:#F5EFE6;--card:#FDFBF8;--dark:#2A1F14;--muted:#9A8570;}}
body{{font-family:'DM Sans',sans-serif;background:var(--cream);max-width:390px;margin:0 auto;overflow-x:hidden;color:var(--dark)}}
.hero-wrap{{position:relative;height:340px;overflow:hidden}}
.hero-bg{{position:absolute;inset:-30px;{hero_bg_style}filter:blur(25px) saturate(1.3) brightness(1.0);transform:scale(1.15);}}
.hero-overlay{{position:absolute;inset:0;background:linear-gradient(180deg,rgba(245,239,230,.68) 0%,rgba(232,213,183,.82) 100%);}}
.status-bar{{position:absolute;top:0;left:0;right:0;z-index:10;display:flex;justify-content:space-between;align-items:center;padding:14px 20px 0;font-size:15px;font-weight:700;color:var(--dark);}}
.hero-logo-wrap{{position:absolute;top:42px;left:20px;z-index:8;width:72px;height:72px;border-radius:14px;overflow:hidden;box-shadow:0 4px 18px rgba(139,115,85,.25);background:var(--latte);display:flex;align-items:center;justify-content:center;}}
.hero-logo-wrap img{{width:100%;height:100%;object-fit:cover}}
.logo-init{{font-family:'DM Serif Display',serif;font-size:22px;color:#fff;}}
.hero-bell{{position:absolute;top:44px;right:20px;z-index:8;width:40px;height:40px;border-radius:20px;background:rgba(255,255,255,.75);display:flex;align-items:center;justify-content:center;font-size:18px;box-shadow:0 2px 10px rgba(139,115,85,.15);}}
.hero-center{{position:absolute;top:50%;left:50%;transform:translate(-50%,-52%);text-align:center;z-index:5;width:230px;}}
.hero-sun{{font-size:18px;color:var(--coffee);margin-bottom:8px}}
.hero-name{{font-family:'DM Serif Display',serif;font-size:26px;color:var(--dark);letter-spacing:1px;text-transform:uppercase;line-height:1.1;}}
.hero-subtitle{{font-size:10px;font-weight:500;letter-spacing:5px;color:var(--coffee);text-transform:uppercase;margin-top:5px;}}
.hero-tagline{{font-family:'Dancing Script',cursive;font-size:16px;color:var(--muted);margin-top:7px;}}
.profile-card{{position:absolute;bottom:18px;left:16px;z-index:8;background:rgba(253,251,248,.92);backdrop-filter:blur(12px);border-radius:16px;padding:13px 18px;min-width:180px;box-shadow:0 8px 30px rgba(139,115,85,.15);}}
.profile-hello{{font-family:'Dancing Script',cursive;font-size:18px;color:var(--coffee)}}
.profile-name{{font-family:'DM Serif Display',serif;font-size:28px;color:var(--dark);line-height:1.1;margin-top:1px;}}
.profile-since{{font-size:12px;color:var(--muted);margin-top:5px}}
.content{{padding:16px 16px 100px}}
.offer-card{{border-radius:18px;overflow:hidden;position:relative;height:220px;margin-bottom:14px;background:linear-gradient(145deg,#3D2B1F 0%,#6B4A30 40%,#8B6845 100%);}}
.offer-body{{position:relative;z-index:2;padding:20px;height:100%;display:flex;flex-direction:column;justify-content:flex-end;}}
.offer-pill{{display:inline-block;background:rgba(196,168,130,.3);border:1px solid rgba(255,255,255,.3);color:#fff;font-size:9px;font-weight:600;letter-spacing:1.8px;text-transform:uppercase;padding:5px 13px;border-radius:20px;margin-bottom:10px;align-self:flex-start;}}
.offer-title-text{{font-family:'DM Serif Display',serif;font-size:26px;color:#fff;line-height:1.1;}}
.offer-italic{{font-family:'Dancing Script',cursive;font-size:26px;color:var(--latte);line-height:1;margin-top:2px;}}
.offer-desc{{font-size:12px;color:rgba(255,255,255,.6);margin-top:7px}}
.offer-badge{{position:absolute;right:16px;bottom:16px;z-index:3;width:76px;height:76px;border-radius:50%;background:rgba(139,115,85,.6);border:2px solid rgba(255,255,255,.3);display:flex;align-items:center;justify-content:center;text-align:center;color:#fff;font-size:8.5px;font-weight:600;letter-spacing:.5px;text-transform:uppercase;line-height:1.25;padding:10px;}}
.actions-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:12px}}
.action-card{{background:var(--card);border-radius:16px;padding:16px 14px 14px;box-shadow:0 2px 12px rgba(139,115,85,.07);text-decoration:none;color:inherit;display:flex;flex-direction:column;gap:3px;position:relative;}}
.action-icon-circle{{width:42px;height:42px;border-radius:50%;background:linear-gradient(135deg,#EFE4D4,#E0CEB0);display:flex;align-items:center;justify-content:center;font-size:18px;margin-bottom:8px;}}
.action-label{{font-size:13px;font-weight:500;color:var(--dark);line-height:1.2}}
.action-desc{{font-size:11px;color:var(--muted);line-height:1.3}}
.action-desc.coffee{{color:var(--coffee)}}
.action-arrow{{position:absolute;right:12px;top:50%;transform:translateY(-50%);color:var(--latte);font-size:20px;}}
.share-banner{{border-radius:18px;overflow:hidden;background:linear-gradient(120deg,#8B7355 0%,#6B5535 100%);padding:20px 18px;display:flex;align-items:center;gap:14px;text-decoration:none;color:inherit;min-height:96px;}}
.share-env{{width:50px;height:50px;flex-shrink:0;border-radius:50%;background:rgba(255,255,255,.15);display:flex;align-items:center;justify-content:center;font-size:22px;}}
.share-text{{flex:1}}
.share-title{{font-family:'Dancing Script',cursive;font-size:22px;color:#fff;font-weight:700;}}
.share-sub{{font-size:11px;color:rgba(255,255,255,.65);margin-top:4px;line-height:1.35}}
.share-arrow{{color:rgba(255,255,255,.45);font-size:22px}}
.bottom-nav{{position:fixed;bottom:0;left:50%;transform:translateX(-50%);width:100%;max-width:390px;background:rgba(245,239,230,.97);backdrop-filter:blur(16px);border-top:1px solid rgba(196,168,130,.2);padding:10px 0 24px;display:flex;z-index:20;}}
.nav-tab{{flex:1;display:flex;flex-direction:column;align-items:center;gap:3px;color:var(--muted);font-size:12px;text-decoration:none;}}
.nav-tab.active{{color:var(--coffee)}}
.nav-icon{{font-size:20px;line-height:1}}
</style>
</head>
<body>
<div class="hero-wrap">
  <div class="hero-bg"></div><div class="hero-overlay"></div>
  <div class="status-bar"><span>9:41</span><span>▪▪▪ 🔋</span></div>
  <div class="hero-logo-wrap">{logo_img}</div>
  <div class="hero-bell">🔔</div>
  <div class="hero-center">
    <div class="hero-sun">☕</div>
    <div class="hero-name">{h_name}</div>
    <div class="hero-subtitle">{d['content']['subtitle']}</div>
    <div class="hero-tagline">{d['tagline']}</div>
  </div>
  <div class="profile-card">
    <div class="profile-hello">Bonjour ♡</div>
    <div class="profile-name">Marie</div>
    <div class="profile-since">☕ Membre depuis {CURRENT_MONTH}</div>
  </div>
</div>
<div class="content">
  <div class="offer-card">
    <div class="offer-body">
      <div class="offer-pill">Ritual du mois</div>
      <div class="offer-title-text">{d['offer_title']}</div>
      <div class="offer-italic">offert ♡</div>
      <div class="offer-desc">{d['offer_sub']}</div>
    </div>
    <div class="offer-badge">RÉSERVÉ<br>À NOS<br>MEMBRES</div>
  </div>
  <div class="actions-grid">
    <a href="#" class="action-card"><div class="action-icon-circle">☕</div><div class="action-label">L'offre du moment</div><div class="action-desc">Votre ritual exclusif</div><span class="action-arrow">›</span></a>
    <a href="{d.get('website') or '#'}" class="action-card"><div class="action-icon-circle">📅</div><div class="action-label">Commander</div><div class="action-desc">Passer votre commande</div><span class="action-arrow">›</span></a>
    <a href="{d.get('instagram_url') or '#'}" class="action-card"><div class="action-icon-circle">📸</div><div class="action-label">Instagram</div><div class="action-desc">Suivez notre univers</div><div class="action-desc coffee">{insta_handle}</div><span class="action-arrow">›</span></a>
    <a href="{d.get('review_url') or '#'}" class="action-card"><div class="action-icon-circle">⭐</div><div class="action-label">Avis Google</div><div class="action-desc">Partagez votre avis</div><span class="action-arrow">›</span></a>
  </div>
  <a href="#" class="share-banner">
    <div class="share-env">💌</div>
    <div class="share-text"><div class="share-title">{d['content']['share_label']}</div><div class="share-sub">{d['content']['share_sub']}</div></div>
    <span class="share-arrow">›</span>
  </a>
</div>
<div class="bottom-nav">
  <a href="#" class="nav-tab active"><span class="nav-icon">☕</span>Club</a>
  <a href="#" class="nav-tab"><span class="nav-icon">👤</span>Compte</a>
</div>
</body>
</html>"""


WALLET_RENDERERS = {
    "brunch": render_wallet_brunch,
    "gastro": render_wallet_gastro,
    "coffee": render_wallet_coffee,
}


# ── DM Mockup ─────────────────────────────────────────────────────────────────

def render_dm_mockup(d: dict, wallet_b64: str) -> str:
    bg_colors = {
        "brunch": "linear-gradient(145deg, #F9E4D4 0%, #F0C8D8 50%, #E8D0C8 100%)",
        "gastro": "linear-gradient(145deg, #1A1200 0%, #2D2000 50%, #1A1A0A 100%)",
        "coffee": "linear-gradient(145deg, #F0E6D6 0%, #E0CEB0 50%, #D4BF9A 100%)",
    }
    bg = bg_colors.get(d["style"], bg_colors["brunch"])
    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=1080, initial-scale=1.0">
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{width:1080px;height:1920px;overflow:hidden;background:{bg};display:flex;flex-direction:column;align-items:center;justify-content:center;font-family:-apple-system,BlinkMacSystemFont,'Helvetica Neue',sans-serif;}}
.ambient{{position:absolute;inset:0;opacity:.4;background:radial-gradient(ellipse 80% 60% at 50% 30%,rgba(255,255,255,.3),transparent)}}
.phone{{position:relative;z-index:1;width:390px;height:844px;background:#fff;border-radius:55px;box-shadow:0 30px 120px rgba(0,0,0,.35),0 0 0 2px rgba(0,0,0,.08),inset 0 0 0 1px rgba(255,255,255,.4);overflow:hidden;}}
.phone::before{{content:'';position:absolute;top:0;left:50%;transform:translateX(-50%);width:126px;height:37px;background:#000;border-radius:0 0 20px 20px;z-index:10;}}
.phone-screen{{width:100%;height:100%;overflow:hidden;border-radius:55px;}}
.phone-screen img{{width:100%;height:100%;object-fit:cover;object-position:top}}
.caption{{position:relative;z-index:1;margin-top:40px;text-align:center;color:rgba(255,255,255,.9);}}
.caption .pretitle{{font-size:16px;letter-spacing:3px;text-transform:uppercase;opacity:.7;margin-bottom:8px}}
.caption .title{{font-size:28px;font-weight:300;letter-spacing:1px}}
.caption .sub{{font-size:14px;margin-top:6px;opacity:.6;letter-spacing:.5px}}
.logo-fidelavis{{position:absolute;bottom:60px;left:50%;transform:translateX(-50%);background:rgba(255,255,255,.15);backdrop-filter:blur(8px);border:1px solid rgba(255,255,255,.25);border-radius:30px;padding:10px 24px;color:#fff;font-size:13px;letter-spacing:2px;text-transform:uppercase;z-index:1;}}
</style>
</head>
<body>
<div class="ambient"></div>
<div class="phone"><div class="phone-screen"><img src="data:image/png;base64,{wallet_b64}" alt="wallet"></div></div>
<div class="caption">
  <div class="pretitle">Votre expérience Fidelavis</div>
  <div class="title">{d['name']}</div>
  <div class="sub">Wallet fidélité personnalisé • {CURRENT_MONTH}</div>
</div>
<div class="logo-fidelavis">✦ Fidelavis</div>
</body>
</html>"""


# ── DM Message ────────────────────────────────────────────────────────────────

DM_TEMPLATES = {
    "brunch": """\
Salut 👋

En découvrant votre univers Instagram, on a imaginé une mini expérience Fidelavis spécialement pour {name} ✨

On vous a préparé une démo personnalisée juste ici 👇
{demo_url}

C'est gratuit, sans engagement — juste pour vous montrer ce que ça peut donner 🌸

Belle journée !
L'équipe Fidelavis
""",
    "gastro": """\
Bonjour,

En découvrant l'univers de {name}, nous avons imaginé une expérience Fidelavis sur-mesure pour votre établissement.

Votre démo personnalisée est disponible ici :
{demo_url}

Nous serions ravis d'échanger sur la façon dont Fidelavis peut valoriser votre clientèle fidèle.

Bien cordialement,
L'équipe Fidelavis
""",
    "coffee": """\
Hello ☕

On a créé une expérience Fidelavis pensée spécialement pour {name} — dans l'esprit de votre univers.

Votre démo ici 👇
{demo_url}

Curieux(se) d'en savoir plus ? On est disponibles !

À bientôt,
L'équipe Fidelavis
""",
}


def generate_dm_message(style: str, name: str, demo_url: str) -> str:
    tpl = DM_TEMPLATES.get(style, DM_TEMPLATES["brunch"])
    return tpl.format(name=name, demo_url=demo_url)


# ── Screenshot via Playwright ─────────────────────────────────────────────────

def screenshot_html(html_path: Path, out_path: Path, width: int = 390, height: int = 844):
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": width, "height": height})
            page.goto(f"file://{html_path.resolve()}")
            page.wait_for_timeout(1400)
            page.screenshot(path=str(out_path), full_page=False)
            browser.close()
        return True
    except Exception as e:
        print(f"  ⚠  screenshot échoué : {e}")
        return False


# ── Main processor ────────────────────────────────────────────────────────────

def process_restaurant(row: dict) -> dict:
    name = (row.get("Name") or "").strip()
    if not name:
        return {}

    cat1 = row.get("First_category") or ""
    cat2 = row.get("Second_category") or ""
    style = detect_style(name, cat1, cat2)
    content = CONTENT[style]

    slug = slugify(name)
    demo_url = f"{DEMO_BASE_URL}/{slug}"
    folder = OUTPUT_DIR / slug
    folder.mkdir(parents=True, exist_ok=True)

    logo_b64 = fetch_url_b64(row.get("Logo_URL"))

    instagram_url = row.get("Instagram_URL") or ""
    print(f"       → scraping Instagram hero…")
    hero_b64 = scrape_instagram_hero(instagram_url)
    if hero_b64:
        print(f"       ✓ hero image récupérée")
    else:
        print(f"       ⚠ pas de hero Instagram — fallback logo/gradient")

    offer_title, offer_sub = content["offers"][hash(name) % len(content["offers"])]
    tagline = content["taglines"][hash(name) % len(content["taglines"])]

    data = {
        "name": name,
        "slug": slug,
        "style": style,
        "template": f"{style}_v1",
        "tagline": tagline,
        "offer_title": offer_title,
        "offer_sub": offer_sub,
        "content": content,
        "instagram_url": instagram_url,
        "review_url": row.get("Review_URL") or "",
        "website": row.get("Website") or "",
        "initials": initials(name),
        "logo_b64": logo_b64,
        "hero_b64": hero_b64,
        "demo_url": demo_url,
    }

    wallet_html = WALLET_RENDERERS[style](data)
    wallet_html_path = folder / "wallet.html"
    wallet_html_path.write_text(wallet_html, encoding="utf-8")

    wallet_png_path = folder / "wallet.png"
    wallet_ok = screenshot_html(wallet_html_path, wallet_png_path, 390, 844)

    if wallet_ok and wallet_png_path.exists():
        wallet_b64_for_mockup = base64.b64encode(wallet_png_path.read_bytes()).decode()
        mockup_html = render_dm_mockup(data, wallet_b64_for_mockup)
        mockup_html_path = folder / "dm_mockup.html"
        mockup_html_path.write_text(mockup_html, encoding="utf-8")
        screenshot_html(mockup_html_path, folder / "dm_mockup.jpg", 1080, 1920)

    msg = generate_dm_message(style, name, demo_url)
    (folder / "message.txt").write_text(msg, encoding="utf-8")

    infos = {
        "name": name,
        "slug": slug,
        "style": style,
        "template": data["template"],
        "address": row.get("Full_Address") or "",
        "phone": row.get("Phone_Standard_format") or "",
        "email": row.get("Email_From_WEBSITE") or "",
        "website": row.get("Website") or "",
        "instagram_url": instagram_url,
        "review_url": row.get("Review_URL") or "",
        "nom_gerant": row.get("nom_gerant") or "",
        "rating": row.get("Average_rating"),
        "reviews_count": row.get("Reviews_count"),
        "premium_score": row.get("premium_score"),
        "statut_contact": row.get("statut_contact") or "non_contacte",
        "ambassadeur": row.get("ambassadeur") or "",
        "hero_scraped": bool(hero_b64),
        "demo_url": demo_url,
        "generated_at": date.today().isoformat(),
        "offer": offer_title,
        "tagline": tagline,
    }
    (folder / "infos.json").write_text(json.dumps(infos, ensure_ascii=False, indent=2), encoding="utf-8")

    return infos


def read_excel(excel_path: Path) -> list:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row = dict(zip(headers, vals))
        if row.get("Name"):
            rows.append(row)
    wb.close()
    return rows


def update_statut_excel(excel_path: Path, slug: str, statut: str):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    statut_col = headers.index("statut_contact") + 1 if "statut_contact" in headers else None
    name_col = headers.index("Name") + 1 if "Name" in headers else None
    if not statut_col or not name_col:
        print("  ⚠  Colonnes manquantes dans l'Excel")
        return
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value or ""
        if slugify(name) == slug:
            ws.cell(r, statut_col).value = statut
            break
    wb.save(excel_path)
    print(f"  ✓  Statut mis à jour → {statut}")


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_PATH

    if not excel_path.exists():
        print(f"❌  Fichier Excel introuvable : {excel_path}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print(f"\n🌸  Fidelavis Lifestyle Generator v2.1")
    print(f"    Excel  : {excel_path.name}")
    print(f"    Output : {OUTPUT_DIR}\n")

    rows = read_excel(excel_path)
    print(f"→ {len(rows)} restaurants trouvés\n")

    results = []
    for i, row in enumerate(rows, 1):
        name = row.get("Name", "?")
        print(f"[{i:02d}/{len(rows)}] {name}")
        style = detect_style(
            row.get("Name") or "",
            row.get("First_category") or "",
            row.get("Second_category") or "",
        )
        print(f"       style → {style}_v1")
        infos = process_restaurant(row)
        if infos:
            results.append(infos)
            hero_status = "✓ hero Instagram" if infos.get("hero_scraped") else "⚠ fallback"
            print(f"       ✓ {infos['slug']}/ [{hero_status}]")
        print()

    print("─" * 50)
    styles_count = {}
    hero_count = sum(1 for r in results if r.get("hero_scraped"))
    for r in results:
        styles_count[r["style"]] = styles_count.get(r["style"], 0) + 1
    print(f"✅  {len(results)} wallets générés")
    print(f"📸  {hero_count}/{len(results)} avec photo Instagram réelle")
    for style, count in sorted(styles_count.items()):
        print(f"   • {style}_v1 : {count} restaurants")


if __name__ == "__main__":
    main()
