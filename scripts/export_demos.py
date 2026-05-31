#!/usr/bin/env python3
"""
Exporte la liste des restaurants ayant une démo (mockup + page demo + handle
Instagram) au format CSV et XLSX, en respectant l'en-tête utilisé par le
fichier exemple `restaurants.csv` :

  restaurant, instagram, wallet_link, video_link, offer, code_table, status, notes

Sortie :
  dm_videos/restaurants_demo.csv
  dm_videos/restaurants_demo.xlsx

Usage : python3 scripts/export_demos.py
"""
from __future__ import annotations
import csv, json, re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT       = Path(__file__).resolve().parent.parent
OUT_CSV    = ROOT / "dm_videos" / "restaurants_demo.csv"
OUT_XLSX   = ROOT / "dm_videos" / "restaurants_demo.xlsx"
PUBLIC     = "https://app.cartefidelavis.com"

COLUMNS = ["restaurant", "instagram", "wallet_link", "video_link",
           "offer", "code_table", "status", "notes"]

EXCLUDED = {
    "assets", "data", "fidelavis-admin", "admin", "apps-script",
    "images", "scripts", "templates", "screenshots", "mockups",
    "pipeline", "dm_videos", ".github", ".git",
}


def normalize_instagram(url: str) -> str:
    """Nettoie l'URL Instagram : enlève le slash final, normalise www."""
    if not url:
        return ""
    url = url.strip().rstrip("/")
    # Normalise www.instagram.com (forme canonique du fichier exemple)
    url = re.sub(r"^https?://(www\.)?instagram\.com/",
                 "https://www.instagram.com/", url, flags=re.IGNORECASE)
    return url


def collect_demos() -> list[dict]:
    rows = []
    for p in sorted(ROOT.iterdir()):
        if not p.is_dir() or p.name.startswith((".", "_")) or p.name in EXCLUDED:
            continue
        cfg_path  = p / "config.json"
        demo_idx  = p / "demo" / "index.html"
        mockup    = ROOT / "mockups" / f"{p.name}-iphone.png"
        if not (cfg_path.exists() and demo_idx.exists() and mockup.exists()):
            continue
        try:
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        # video_link : URL publique du MP4 généré si présent
        video_mp4 = ROOT / "dm_videos" / f"{p.name}-dm.mp4"
        video_link = f"{PUBLIC}/dm_videos/{p.name}-dm.mp4" if video_mp4.exists() else ""

        rows.append({
            "restaurant":  cfg.get("name") or p.name,
            "instagram":   normalize_instagram(cfg.get("instagramUrl", "")),
            "wallet_link": f"{PUBLIC}/{p.name}/demo/",
            "video_link":  video_link,
            "offer":       (cfg.get("activeCoupon") or {}).get("title", "").strip(),
            "code_table":  "",
            "status":      "ready" if video_link else "pending_video",
            "notes":       "",
        })
    return rows


def write_csv(rows: list[dict]) -> Path:
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS, quoting=csv.QUOTE_MINIMAL)
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return OUT_CSV


def write_xlsx(rows: list[dict]) -> Path:
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Démos restaurants"

    # Styles
    header_fill   = PatternFill(start_color="1A1A1C", end_color="1A1A1C", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF", size=12)
    body_font     = Font(size=11)
    ready_fill    = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ready_font    = Font(bold=True, color="166534", size=11)
    pending_fill  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    pending_font  = Font(bold=True, color="92400E", size=11)
    border_thin   = Border(left=Side(style="thin", color="E5E5E5"),
                           right=Side(style="thin", color="E5E5E5"),
                           top=Side(style="thin", color="E5E5E5"),
                           bottom=Side(style="thin", color="E5E5E5"))
    align_left    = Alignment(horizontal="left", vertical="center", wrap_text=False)
    align_center  = Alignment(horizontal="center", vertical="center")

    # Header
    for i, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align_center
        c.border = border_thin
    ws.row_dimensions[1].height = 28

    # Data
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            cell.font = body_font
            cell.alignment = align_left
            cell.border = border_thin
            # Status colorisé
            if col == "status" and row[col] == "ready":
                cell.fill = ready_fill
                cell.font = ready_font
                cell.alignment = align_center
            elif col == "status" and row[col] == "pending_video":
                cell.fill = pending_fill
                cell.font = pending_font
                cell.alignment = align_center
            # Liens cliquables
            if col in ("instagram", "wallet_link", "video_link") and row[col]:
                cell.hyperlink = row[col]
                cell.font = Font(color="2563EB", underline="single", size=11)

    # Largeurs auto
    widths = {"restaurant": 34, "instagram": 42, "wallet_link": 46,
              "video_link": 42, "offer": 38, "code_table": 12,
              "status": 12, "notes": 28}
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[col]

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    return OUT_XLSX


def main():
    rows = collect_demos()
    if not rows:
        raise SystemExit("Aucun resto avec démo + mockup trouvé.")
    write_csv(rows)
    write_xlsx(rows)
    print(f"✓ {len(rows)} restos exportés :")
    print(f"   • {OUT_CSV.relative_to(ROOT)}")
    print(f"   • {OUT_XLSX.relative_to(ROOT)}")
    print()
    for r in rows:
        print(f"   - {r['restaurant']:35s}  @{r['instagram'].split('/')[-1]:25s}  {r['offer']}")


if __name__ == "__main__":
    main()
